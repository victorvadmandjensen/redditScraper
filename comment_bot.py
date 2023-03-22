import praw
from pathlib import Path
import pathlib
from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter
# import login.py file
import login
import datetime
from bs4 import BeautifulSoup

# Get Reddit instance from login.py
reddit = login.get_reddit()

# Define subreddit to scrape and number of posts
#subreddit = "botsRights"
#post_number = 1000


# Create empty list of data and populate it
#data_list = []
#for submission in reddit.subreddit(subreddit).top(time_filter = "all", limit=None):
 #   data_list.append(submission)

# print(data_list)

# Get the path of the xlsx file for data to be saved in
a = pathlib.Path(__file__)
p = a.parents[0] / "python_comment_data.xlsx"
data_file = Path(p)

post_file = a.parents[0] / "python_post_data.xlsx"


# Check if Excel file exists, and if not create it
if data_file.is_file():
    print("Data file exists")
else:
    workbook = xlsxwriter.Workbook("python_comment_data.xlsx")
    print("Data file has been created")

# Define workbook to write to and as well as load post title workbook
wb = Workbook()
ws5 = wb.create_sheet(title="Post comments")
ws5.cell(column=2, row=1).value = "Post comment"
ws5.cell(column=1, row=1).value = "Post ID"

wb2 = load_workbook("python_post_data.xlsx")
ws_titles = wb2["Post IDs"]

# create counter so we can keep track of rows, and start on 1, so when we add we go to row 2
counter = 1
# loop through the specific rows in the relevant sheet . Set to 1005 just to make sure, but we do not have that many posts!
for i in range(1, 50):
    # read the ID of the post we are currently at
    current_post = ws_titles.cell(column=1,row=i+1).value
    # if current_post is empty in Excel (as in post has been deleted) just continue
    if current_post == None:
        continue

    find_post = reddit.submission(current_post)
    # The below code traverses the tree of comments breadth-first, meaning it looks for all comments on one level, before moving to tbe next level
    # Good explanation for breadth-first: https://old.reddit.com/r/explainlikeimfive/comments/1g7j4t/eli5_whats_the_difference_between_depth_first/cahhrn1/ 
    find_post.comments.replace_more(limit=None)
    for comment in find_post.comments.list():
        # increment counter
        counter = counter + 1
        # put data into worksheet
        ws5.cell(column=1, row=counter).value = str(find_post)
        ws5.cell(column=2,row=counter).value = ''.join(BeautifulSoup(comment.body_html, "html.parser").findAll(string=True) )
        
# Save the workbook
wb.save(filename=data_file)