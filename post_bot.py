import praw
from pathlib import Path
import pathlib
from openpyxl import Workbook
import xlsxwriter
# import login.py file
import login
import datetime
from bs4 import BeautifulSoup

# Get Reddit instance from login.py
reddit = login.get_reddit()

# Define subreddit to scrape and number of posts
subreddit = "botsRights"
post_number = 1000


# Create empty list of data and populate it
data_list = []
for submission in reddit.subreddit(subreddit).top(time_filter = "all", limit= 200):
    data_list.append(submission)

# print(data_list)

# Get the path of the xlsx file for data to be saved in
a = pathlib.Path(__file__)
p = a.parents[0] / "python_data.xlsx"
data_file = Path(p)

# Check if Excel file exists, and if not create it
if data_file.is_file():
    print("Data file exists")
else:
    workbook = xlsxwriter.Workbook("python_data.xlsx")
    print("Data file has been created")

# Define workbook to write to and file and add column names
wb = Workbook()
ws1 = wb.create_sheet(title="Post titles and IDs")
ws1.cell(column=1, row=1).value = "Post title"
ws1.cell(column=2, row=1).value = "Post ID"

ws2 = wb.create_sheet(title="Post times")
ws2.cell(column=1, row=1).value = "Post time"

ws3 = wb.create_sheet(title="Post authors")
ws3.cell(column=1, row=1).value = "Post author"

#ws4 = wb.create_sheet(title="Post selftexts")
#ws4.cell(column=1, row=1).value = "Post selftexts")

ws5 = wb.create_sheet(title="Post comments")
ws5.cell(column=2, row=1).value = "Post comment"
ws5.cell(column=1, row=1).value = "Post ID"

number_comments = 0

# Iterate over items in data_list and add the title to the Excel sheet
for data_item in range(1, len(data_list)):
    # Check if the author is deleted
    if not data_list[data_item].author:
        continue
    ws1.cell(column=1, row=data_item+1).value = data_list[data_item].title
    ws1.cell(column=2, row=data_item+1).value = data_list[data_item].id
    # convert Unix time of posts to UTC
    new_time = datetime.datetime.utcfromtimestamp(data_list[data_item].created_utc)
    ws2.cell(column=1, row=data_item+1).value = new_time
    ws3.cell(column=1, row=data_item+1).value = data_list[data_item].author.name
    #ws4.cell(column=1, row=data_item).value = data_list[data_item].selftext
    #ws5.cell(column=1, row=data_item+1).value = data_list[data_item].id
    #for comment in data_list[data_item].comments:
     #   number_comments += 1
      #  ws5.cell(column=2,row=data_item+1).value = ''.join(BeautifulSoup(comment.body_html, "html.parser").findAll(string=True))
       # if number_comments > 1000:
        #    break

#print number of submissions
print(f"Number of submissions is: {len(data_list)}")

# print number of comments
print(f"Number of comments is: {number_comments}, but this is not calculated correctly.")

# Save the workbook
wb.save(filename=data_file)